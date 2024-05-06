import json
from datetime import datetime
from io import BytesIO

import xlsxwriter
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import CreateView, UpdateView, DeleteView, TemplateView
from django.views.generic.base import View
from openpyxl import load_workbook
from django.db.models import Q
from core.pos.forms import ProductForm, Product, Category
from core.security.mixins import PermissionMixin, ModuleMixin

class ProductListView(PermissionMixin, TemplateView):
    template_name = 'scm/product/list.html'
    permission_required = 'view_product'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST.get('action')
        try:
            if action == 'search':
                data = []
                for p in Product.objects.all():
                    data.append(p.toJSON())
            elif action == 'upload_excel':
                archive = request.FILES.get('archive')
                if archive:
                    print("Archivo recibido:", archive.name)
                    with transaction.atomic():
                        workbook = load_workbook(filename=archive, data_only=True)
                        excel = workbook.active
                        for row in range(2, excel.max_row + 1):
                            name = excel.cell(row=row, column=2).value
                            description = excel.cell(row=row, column=3).value
                            bar = excel.cell(row=row, column=4).value
                            category = excel.cell(row=row, column=5).value
                            price = float(excel.cell(row=row, column=6).value)
                            pvp  = float(excel.cell(row=row, column=7).value)
                            stock  = float(excel.cell(row=row, column=8).value)
                            unit  = excel.cell(row=row, column=9).value
                            
                            if unit == 'unidad' or unit=='Unidad' or unit=='UNIDAD':
                                unit = 'unidad'
                            elif unit == 'kg' or unit == 'kilogramo' or unit=='KG' or unit == 'Kilogramo' or unit == 'Kg':
                                unit = 'kilogramo'
                                
                            # Validar que el nombre de la categoría no esté vacío
                            if category:
                                category, _ = Category.objects.get_or_create(name=category)
                                print('ID DE LA CATEGORIA: ', category.id)
                                existing_product = Product.objects.filter(name=name, codebar=bar).first()
                                print(existing_product)
                                if existing_product:
                                    # print("stock exist", existing_product.stock)
                                    print('prodcuto existente', existing_product.name)
                                    existing_product.name = name
                                    existing_product.desc = description
                                    existing_product.bar = bar
                                    existing_product.category = category
                                    existing_product.price = price
                                    existing_product.pvp = pvp
                                    existing_product.stock = float(existing_product.stock) + stock
                                    existing_product.unit = unit
                                    existing_product.save()
                                else:
                                    print('se va a crear')
                                    product = Product.objects.create(
                                        name=name,
                                        desc=description,
                                        codebar=bar,
                                        category=category, 
                                        price=price,
                                        pvp=pvp,
                                        stock=stock,
                                        unit=unit,
                                    )
                                    print('producto nuevo', product.name)
                            else:
                                # Manejar el caso cuando el nombre de la categoría está ausente
                                print('El nombre de la categoría está ausente en el archivo Excel.')
                                break
                    data['success'] = 'Datos cargados correctamente desde el archivo Excel.'
                else:
                    data['error'] = 'No se ha proporcionado ningún archivo para subir.'
            else:
                data['error'] = 'Acción no válida.'
        except Exception as e:
            print("Error:", e)  # Agregar instrucción de depuración
            data['error'] = str(e)

        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['create_url'] = reverse_lazy('product_create')
        context['title'] = 'Listado de Productos'
        return context

class ProductCreateView(PermissionMixin, CreateView):
    model = Product
    template_name = 'scm/product/create.html'
    form_class = ProductForm
    success_url = reverse_lazy('product_list')
    permission_required = 'add_product'

    def validate_data(self):
        data = {'valid': True}
        try:
            name = self.request.POST['name'].strip()
            category = self.request.POST['category']
            if len(category):
                if Product.objects.filter(name__iexact=name, category_id=category):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'add':
                product = Product()
                product.name = request.POST['name']
                product.desc = request.POST['desc']
                product.codebar = request.POST['codebar']
                product.category_id = request.POST['category']
                if 'image' in request.FILES:
                    product.image = request.FILES['image']
                product.pvp = float(request.POST['pvp'])
                if product.category.inventoried:
                    product.price = float(request.POST['price'])
                else:
                    product.price = product.pvp
                product.unit = request.POST['unit']
                print('PROBANDO LO QUE LLEGA', product.unit)
                product.save()
            elif action == 'search_category_id':
                data = Category.objects.get(pk=request.POST['id']).toJSON()
            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['list_url'] = self.success_url
        context['title'] = 'Nuevo registro de un Producto'
        context['action'] = 'add'
        return context


class ProductUpdateView(PermissionMixin, UpdateView):
    model = Product
    template_name = 'scm/product/create.html'
    form_class = ProductForm
    success_url = reverse_lazy('product_list')
    permission_required = 'change_product'

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def validate_data(self):
        data = {'valid': True}
        try:
            id = self.get_object().id
            name = self.request.POST['name'].strip()
            category = self.request.POST['category']
            if len(category):
                if Product.objects.filter(name__iexact=name, category_id=category).exclude(id=id):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'edit':
                product = self.object
                product.name = request.POST['name']
                product.desc = request.POST['desc']
                product.codebar = request.POST['codebar']
                product.category_id = request.POST['category']
                if 'image-clear' in request.POST:
                    product.remove_image()
                if 'image' in request.FILES:
                    product.image = request.FILES['image']
                product.pvp = float(request.POST['pvp'])
                if product.category.inventoried:
                    product.price = float(request.POST['price'])
                else:
                    product.price = product.pvp
                product.unit = request.POST['unit']
                product.save()

            elif action == 'search_category_id':
                data = Category.objects.get(pk=request.POST['id']).toJSON()
            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['list_url'] = self.success_url
        context['title'] = 'Edición de un Producto'
        context['action'] = 'edit'
        return context


class ProductDeleteView(PermissionMixin, DeleteView):
    model = Product
    template_name = 'scm/product/delete.html'
    success_url = reverse_lazy('product_list')
    permission_required = 'delete_product'

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            self.get_object().delete()
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Notificación de eliminación'
        context['list_url'] = self.success_url
        return context


class ProductStockAdjustmentView(ModuleMixin, TemplateView):
    template_name = 'scm/product/stock_adjustment.html'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'search_products':
                data = []
                ids = json.loads(request.POST['ids'])
                term = request.POST['term']
                search = Product.objects.filter(category__inventoried=True).exclude(id__in=ids).order_by('name')
                if len(term):
                    search = search.filter(Q(name__icontains=term) | Q(codebar__icontains=term))
                    search = search[0:10]
                for p in search:
                    item = p.toJSON()
                    item['value'] = '{} / {} / {} / {}'.format(p.name, p.category, p.desc, p.codebar)
                    data.append(item)
            elif action == 'create':
                with transaction.atomic():
                    for p in json.loads(request.POST['products']):
                        product = Product.objects.get(pk=p['id'])
                        product.stock = float(p['newstock'])
                        product.save()
            else:
                data['error'] = 'No ha ingresado una opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Ajuste de Stock de Productos'
        return context


class ProductExportExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        try:
            headers = {
                'Código': 15,
                'Producto': 75,
                'Descripción': 75,
                'Código Barra': 75,
                'Categoría': 20,
                'Precio de Compra': 20,
                'Precio venta': 20,
                'Stock': 10,
                'unit': 30,
            }

            output = BytesIO()
            workbook = xlsxwriter.Workbook(output)
            worksheet = workbook.add_worksheet('productos')
            cell_format = workbook.add_format({'bold': True, 'align': 'center', 'border': 1})
            row_format = workbook.add_format({'align': 'center', 'border': 1})
            index = 0
            for name, width in headers.items():
                worksheet.set_column(first_col=0, last_col=index, width=width)
                worksheet.write(0, index, name, cell_format)
                index += 1
            row = 1
            for product in Product.objects.filter().order_by('id'):
                worksheet.write(row, 0, product.id, row_format)
                worksheet.write(row, 1, product.name, row_format)
                worksheet.write(row, 2, product.desc, row_format)
                worksheet.write(row, 3, product.codebar, row_format)
                worksheet.write(row, 4, product.category.name, row_format)
                worksheet.write(row, 5, format(product.price, '.2f'), row_format)
                worksheet.write(row, 6, format(product.pvp, '.2f'), row_format)
                worksheet.write(row, 7, format(product.stock, '.2f'), row_format)
                worksheet.write(row, 8, product.unit, row_format)
                row += 1
            workbook.close()
            output.seek(0)
            response = HttpResponse(output,
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="PRODUCTOS_{}.xlsx"'.format(
                datetime.now().date().strftime('%d_%m_%Y'))
            return response
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('product_list'))


class GeneradorView(ModuleMixin, TemplateView):
    template_name = 'scm/product/qr.html'

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'search_products':
                data = []
                ids = json.loads(request.POST['ids'])
                term = request.POST['term']
                search = Product.objects.filter(category__inventoried=True).exclude(id__in=ids).order_by('name')
                if len(term):
                    search = search.filter(name__icontains=term)
                    search = search[0:10]
                for p in search:
                    item = p.toJSON()
                    item['value'] = '{} / {}'.format(p.name, p.category.name)
                    data.append(item)
            elif action == 'create':
                with transaction.atomic():
                    for p in json.loads(request.POST['products']):
                        product = Product.objects.get(pk=p['id'])
            else:
                data['error'] = 'No ha ingresado una opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Generador de Productos'
        return context
